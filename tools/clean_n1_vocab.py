#!/usr/bin/env python3
"""Clean N1 vocabulary rows based on the latest audit report."""

from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_DIR / "source" / "日语单词词汇表.xlsx"
REPORT_PATH = PROJECT_DIR / "reports" / "N1-检查明细.json"

SIMPLIFIED_FIXES = {
    "诚实": "誠実",
    "促进": "促進",
    "定义": "定義",
    "发覚": "発覚",
    "发足": "発足",
}


def main() -> None:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = WORKBOOK_PATH.with_name(f"日语单词词汇表-N1清理前备份-{stamp}.xlsx")
    shutil.copy2(WORKBOOK_PATH, backup_path)

    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    exact_rows = {int(row[0]) for row in report["exact_duplicates"]}

    internal_groups: dict[int, list[int]] = {}
    for row in report["internal_duplicates"]:
        internal_groups.setdefault(int(row[0]), []).append(int(row[1]))

    internal_delete_rows = set()
    for rows in internal_groups.values():
        remaining = [row for row in sorted(rows) if row not in exact_rows]
        internal_delete_rows.update(remaining[1:])

    rows_to_delete = exact_rows | internal_delete_rows

    workbook = load_workbook(WORKBOOK_PATH)
    if "N1" not in workbook.sheetnames:
        raise ValueError("Excel 中找不到 N1 工作表")
    sheet = workbook["N1"]

    simplified_fixed = []
    for row in range(2, sheet.max_row + 1):
        value = str(sheet.cell(row=row, column=2).value or "").strip()
        if value in SIMPLIFIED_FIXES:
            sheet.cell(row=row, column=2).value = SIMPLIFIED_FIXES[value]
            simplified_fixed.append((row, value, SIMPLIFIED_FIXES[value]))

    trailing_start = None
    for row in range(2, sheet.max_row + 1):
        level = str(sheet.cell(row=row, column=1).value or "").strip().lower()
        rest_blank = all(
            not str(sheet.cell(row=row, column=col).value or "").strip()
            for col in (2, 3, 4)
        )
        if level == "n1" and rest_blank:
            trailing_start = row
            break

    if trailing_start:
        sheet.delete_rows(trailing_start, sheet.max_row - trailing_start + 1)
        rows_to_delete = {row for row in rows_to_delete if row < trailing_start}

    for row in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row, 1)

    workbook.save(WORKBOOK_PATH)

    summary = {
        "backup": str(backup_path),
        "deleted_exact_duplicate_rows": len(exact_rows),
        "deleted_internal_duplicate_rows": len(internal_delete_rows),
        "deleted_total_rows": len(rows_to_delete),
        "simplified_fixed": simplified_fixed,
        "trailing_blank_rows_removed_from": trailing_start,
    }
    out_path = PROJECT_DIR / "reports" / "N1-清理记录.json"
    out_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
